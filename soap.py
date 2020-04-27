from typing import List, Any, Tuple
import openpyxl 
import xlsxwriter 
import pandas as pd
import xlrd
import glob, os
import subprocess as sp
from os import path

# represents a single test tube in the test tube rack
class testTube:

    # create a testTube
    def __init__(self, maxConcentration: float, rawMaterialNumber, name: str, pricePerPound: float, concentration: float) -> None:
        # the maximum concentration, price per lb., and concentration cannot be less than 0
        if maxConcentration < 0:
            raise Exception("the formula is not allowed to have a maxconcentration below 0")
        if pricePerPound < 0:
            raise Exception("the formula is not allowed to have a pricePerPound below 0")
        if concentration < 0:
            raise Exception("the formula is not allowed to have a concentration below 0")
        # raw material number cannot be able to be manipulated as an int or float
        if type(rawMaterialNumber) is int:
            rawMaterialNumber = str(rawMaterialNumber)
        # set the name, raw material number, max concentration, price per pound, & concentration for the test tube
        # NOTE: the current implimentation requires that every material have a raw material number: deal with it chemical engineers!
        self.name = name 
        self.rawMaterialNumber = rawMaterialNumber
        self.maxConcentration = maxConcentration 
        self.pricePerPound = pricePerPound 
        self.concentration = concentration 



    # returns the name of the chemical in the test tube
    def __repr__(self) -> str:
        return self.name



    # increase the concentration of the tube by a specified amount
    # returns a bool to indicate that the process was successful
    def addByConcentration(self, amount: float) -> bool:
        # specified amount must be greater than 0
        if amount + self.concentration > self.maxConcentration or amount < 0:
            return False
        self.concentration += amount
        return True



    # decrease the concentration of the tube by specified amount
    # returns a bool to indicate that the process was successful
    def reduceByConcentration(self, amount: float) -> bool:
        # specified amount must be greater than 0
        if self.concentration - amount < 0 or amount < 0: 
            return False 
        self.concentration -= amount
        return True



    # Set the concentration of the tube by replacing with amount
    # returns a bool to indicate that the process was successful
    def setCapacity(self, amount: float) -> bool:
        # checking input bounds, this is using the max concentration attribute, but it could just be 100
        if amount > self.maxConcentration or amount < 0: 
            return False 
        # if the conditions are acceptable, set the current concentration to the input "amount" and return true
        self.concentration = amount 
        return True 



    # get the cost of the tube = price/lb * concentration in the tube
    def getCost(self) -> float:
        # Calculate and return the cost of the chemical in the mixture, based on concentration and price per pound
        return (self.pricePerPound * self.concentration) / 100 



    # reduce the concentration of the tube by this $amount
    def reduceByPrice(self, amount: float) -> bool:
        if self.pricePerPound != 0:
            # get the cost of the chemical using the getCost method
            cost = self.getCost() 
            # calculate the new desired cost based on the amount and current cost
            newcost = cost - amount 
            # check to make sure the cost is not out of bounds
            if newcost < 0 or amount < 0:
                # necessary?...probably 
                return False 
            # calculate the new concentration
            self.concentration = 100 * newcost / self.pricePerPound 
            return True
        else:
            return True



    # increase the concentration of the tube by this $amount
    def addByPrice(self, amount: float) -> bool:
        if self.pricePerPound !=0:
            # calculate a new cost based off of the amount and current cost
            newcost = self.getCost() + amount 
            # calculate the new concentration based on the newcost
            increasedConcentration = 100 * newcost / self.pricePerPound
            # check that the concentration does not go out of bound 
            if increasedConcentration > 100: 
                return False
            # set the new concentration
            self.concentration = increasedConcentration 
            return True
        else:
            return True



    def concentrationIfFillToPrice(self, amount: float) -> float:
        if self.pricePerPound != 0:
            # calculate the new cost based on the current cost and input amount
            newcost = self.getCost() + amount 
            # calculate the new concentration
            increasedConcentration = 100 * newcost / self.pricePerPound 
            # return the increased concentration to be used later
            return increasedConcentration 
        else:
            return self.concentration



    # Set the concentration of the tube to 0
    def emptyTube(self) -> None:
        # set the current concentration equal to 0
        self.concentration = 0 



###################################################################################################################################

class testTubeRack:

    # test tube racks are objects which contain multiple test tubes
    def __init__(self, name: str, pricePoint: float) -> None:
        if pricePoint < 0:
            raise Exception("the formula is not allowed to have a pricepoint below 0")
        # set the pricepoint argument
        self.pricePoint = pricePoint 
        # add the testtube to the rack of tubes
        self.testTubes = [] 
        # associate the name in the rack
        self.name = name 
        self.notes = ""



    # changes the order of ingredients in the test tubes
    def swapIngredients(self, ingredientOne: str, ingredientTwo: str) -> None:
        indexCounter = 0
        # search through the test tubes
        for tt in self.testTubes:
            # find the position of the first ingredient
            if tt.name == ingredientOne:
                ingredientOneIndex = indexCounter
            # find the position of the second ingredient
            if tt.name == ingredientTwo:
                ingredientTwoIndex = indexCounter
            indexCounter += 1
        # create temporary test tube variables
        tempTestTubeOne = self.testTubes[ingredientOneIndex]
        tempTestTubeTwo = self.testTubes[ingredientTwoIndex]
        # use the temporary test tube variables to switch the positions of the two test tubes
        self.testTubes[ingredientOneIndex] = tempTestTubeTwo
        self.testTubes[ingredientTwoIndex] = tempTestTubeOne



    # name of the testTubeRack
    def __repr__(self) -> str:
        # returns the name of the formula
        return self.name 



    # Report the formula name and pricepoint, then the ingredients and their corresponding concentrations
    # Gives a high level overview of the formula
    def formula(self) -> List[Any]:
        # initialize the formula variable
        a = [] 
        # first array will be the name of the formula, and the pricepoint of the formula
        a.append([self.name, self.pricePoint]) 
        # the remaining arrays are the names of the testtubes and their corresponding concentrations in the formula
        for i in range(0, len(self.testTubes)):  
            a.append([self.testTubes[i].name, self.testTubes[i].concentration])
        return a



    # create a list out of the test tubes in the test tube rack
    # useful for writing .xlsx sheets
    def createIngredientArray(self) -> List[Any]:
        finalList = []
        # records the important information for each test tube and adds it to the final list
        for t in self.testTubes:
            pNum = t.rawMaterialNumber
            desc = t.name
            quantity = t.concentration
            curCost = t.pricePerPound
            partialArray = [pNum, desc, quantity, curCost]
            finalList.append(partialArray)
        # returns the list of test tubes
        return finalList



    # writes a .xlsx file based on the ingredients in the formula
    def exportFormula(self) -> None:
        # creates workbook
        workbook = xlsxwriter.Workbook(self.name + '.xlsx')
        # creates worksheet 
        worksheet = workbook.add_worksheet(self.name)
        # creates important labels 
        worksheet.write('B1', 'Part Number') 
        worksheet.write('C1', 'Description') 
        worksheet.write('E1', 'Quantity') 
        worksheet.write('F1', 'Current Cost')
        # gets the data we want to write to the worksheet 
        scores = self.createIngredientArray()
        # Start from the first cell. Rows and columns are zero indexed. 
        row = 1
        col = 1
        # Iterate over the test tubes and write it out row by row. 
        for partNumber, Description, Quantity, currentCost in (scores):
            # write rmn 
            worksheet.write(row, col, partNumber)
            # write name 
            worksheet.write(row, col + 1, Description)
            # write concentration
            worksheet.write(row, col + 3, Quantity)
            # write current cost
            worksheet.write(row, col + 4, currentCost)
            row += 1
        # write price point
        worksheet.write(0,8, self.pricePoint)
        # write notes
        worksheet.write(2,8, self.notes)
        # close the workbook
        workbook.close()  
        


    # change the target price of the rack
    # returns bool to indicate whether the request was successful
    def changePricePoint(self, amount: float) -> bool:
        # make sure amount is not a str
        res = isinstance(amount, str)
        if res == True:
            return False
        # check bounds
        if amount < 0: 
            return False
        # change the pricepoint to the input amount 
        self.pricePoint = amount     
        return True



    # create a new tube, name, price/lb, and desired concentration
    def createRackTube(self, rawMaterialNumber, name: str, pricePerPound: float, concentration: float) -> bool:
        # check to see if the new concentration will push the total rack concentration over 100/ out of bounds
        if concentration <=  self.unusedRackConcentration(): 
            # add new chemical to the formula
            self.testTubes.append(testTube(100,rawMaterialNumber,name,pricePerPound,concentration)) 
            return True
        return False



    # returns the total cost of the rack by suming the cost * concentration of each ingredient
    def getCost(self) -> float:
        # initialize the total cost
        totalCost = 0 
        # loop through the length of the testtubes
        for i in range(0, len(self.testTubes)): 
            # calculate the total cost of the formula by adding the cost of each ingredient
            totalCost += self.testTubes[i].getCost() 
        return totalCost



    # returns the sum of the concentration of all ingredients in the rack
    def sumRackConcentration(self) -> float:
        # initialize totalConcentrations
        totalConcentration = 0 
        for i in range(0, len(self.testTubes)):
            # Calculate the total concentration by adding each individual concentrations
            totalConcentration += self.testTubes[i].concentration 
        return totalConcentration



    # return 100 - total Rack concentration. How much room is left to add ingredients?
    def unusedRackConcentration(self) -> float:
        # check the available concentration in the rack
        return 100 - self.sumRackConcentration() 



    # if the cost of the ingredients is greater than the pricepoint, decrease the concentration of each listed ingredient until 
    # the pricpoint is reached. If an ingrdient reaches 0% concentration and the pricepoint is not reached, continue removing 
    # the concentration evenly from the other listed ingredients.
    def reduceToPrice(self, reduceableIngredients: List[str]) -> None:
        # calculate the new desired price of the formula
        reduceCurrentPriceBy = self.getCost() - self.pricePoint 
        # make sure input is reasonable and not out of bounds
        if reduceCurrentPriceBy > 0: 
            if len(reduceableIngredients) <= 0:
                return
            # this is important for breaking the while loop, reducecurrentpriceby changes later
            remainingReducableCost = reduceCurrentPriceBy 
            # get their prices in a for loop, initializing this variable here
            ingredientPrices = [] 
            # initializing the lowest cost of all the ingredients, used later
            lowestCost = 0 
            # condition to break the while loop, once all cost is accounted for, loop will be broken
            while remainingReducableCost != 0: 
                 # loop to get the position of the reduceable ingredients in the rack and important information
                for i in range(0,len(self.testTubes)):
                    # finds reduceable ingredients
                    if self.testTubes[i].name in reduceableIngredients: 
                        if self.testTubes[i].getCost() != 0:
                            # returns current cost of each ingredient
                            ingredientPrices.append(self.testTubes[i].getCost()) 
                        else:
                            reduceableIngredients.remove(self.testTubes[i].name)
                            if len(reduceableIngredients) <= 0:
                                return
                if ingredientPrices == []:
                    return
                # determine the lowest cost in the available reduceable ingredients
                lowestCost = min(ingredientPrices) 
                # this is important to know if we can or cannot evenly distribute the change in price to each ingredient
                if remainingReducableCost/len(reduceableIngredients) <= lowestCost: 
                    for i in range(0,len(self.testTubes)):
                        if self.testTubes[i].name in reduceableIngredients:
                            # evenly distribute the change in price to all reduceable ingredients
                            self.testTubes[i].reduceByPrice(remainingReducableCost/len(reduceableIngredients)) 
                    # no need to check remainingreduceablecost while loop is satisfied, the math is good enough
                    break 
                # if they are not evenly reduceable, first reduce everything by the lowest ingredient and remove it from the reduceable ingredients list
                else: 
                    for i in range(0, len(self.testTubes)):
                        # remove any ingredient with the same lowest cost
                        if self.testTubes[i].name in reduceableIngredients and self.testTubes[i].getCost() == lowestCost: 
                            # reduce all ingredients by the lowest price
                            self.testTubes[i].reduceByPrice(lowestCost) 
                            # necessary for breaking the while loop
                            remainingReducableCost -= lowestCost 
                            # if it is the lowest cost, remove from available ingredients
                            reduceableIngredients.remove(self.testTubes[i].name) 
                            # also remove the cost in the list of costs
                            ingredientPrices.remove(lowestCost) 
                        # reduce ingredients that are greater than the lowest cost by the lowest costing ingredient
                        elif self.testTubes[i].name in reduceableIngredients: 
                            self.testTubes[i].reduceByPrice(lowestCost)
                            # to break while loop if condition is satisfied
                            remainingReducableCost -= lowestCost 
                    # reset the ingredient prices because prices and ingredients have changed at this point, but still in the while loop. reinitialize
                    ingredientPrices = []



    # if the cost of the ingredients is less than the price point, increase the concentration of the listed ingredients
    # until the price point is met. If the total concentration of the rack needs to exceed 100% to reach the price point 
    # equally increase the price of each ingredient until the rack concentration = 100%
    def fillToPrice(self, increaseableIngredients: List[str]) -> None:
        # calculate the price the formula needs to increase to
        addCurrentPriceBy = self.pricePoint - self.getCost() 
        # initialize concentration increase to later check for going out of bounds of the max concentration of the rack
        concentrationIncrease = 0 
        # initialize the potential new concentration to later check if we are moving out of bounds
        potentialNewConcentration = 0 
        if addCurrentPriceBy > 0:
            for i in range(0,len(self.testTubes)):
                # find increaseable ingredients
                if self.testTubes[i].name in increaseableIngredients: 
                    if len(increaseableIngredients) <= 0:
                        return
                    # determine what the concentration would be if allowed
                    potentialNewConcentration = self.testTubes[i].concentrationIfFillToPrice(addCurrentPriceBy/len(increaseableIngredients)) 
                    # calculate how much this increase would increase the overall concentration
                    concentrationIncrease += potentialNewConcentration - self.testTubes[i].concentration 
                # if the potential concentration greater than the available concentration in the rack
            if concentrationIncrease > self.unusedRackConcentration(): 
                # recalculate a new price to increase everything evenly by to stay within bounds
                addCurrentPriceBy = ((concentrationIncrease-(concentrationIncrease + self.sumRackConcentration() - 100)) / concentrationIncrease) * addCurrentPriceBy 
            for i in range(0,len(self.testTubes)):
                if self.testTubes[i].name in increaseableIngredients:
                    # evenly increase concentration by the available price/number of increaseable ingredients
                    self.testTubes[i].addByPrice(addCurrentPriceBy/len(increaseableIngredients)) 



    # Use the listed ingredients to make the total rack concentration = 100%
    def fillToConcentration(self, increaseableIngredients: List[str]) -> None:
        increaseConcentrationBy = self.unusedRackConcentration()
        # check to make sure inputs are in bounds
        if self.unusedRackConcentration() > 0: 
             for i in range(0,len(self.testTubes)):
                 # find reduceable ingredients in the rack
                if self.testTubes[i].name in increaseableIngredients: 
                    # increase ingredients evenly by available concentration/number of increaseable ingredients
                    self.testTubes[i].addByConcentration(increaseConcentrationBy / len(increaseableIngredients)) 



    # Set listed tubes concentration to 0 by calling empty tube in the testTube object
    def emptyTubes(self, emptyIngredients: List[str]) -> None:
        for i in range(0,len(self.testTubes)):
            # find ingredients in rack
            if self.testTubes[i].name in emptyIngredients: 
                # set concentraiton of the tube to 0
                self.testTubes[i].emptyTube() 



    # Example: Current Tube cost = $4, amount = $10, increase tube concentration by $6
    # if cost will increase concentration greater than unused rack concentration call fill to 100% concentration
    def increaseRackTubeToPrice(self, ingredient: str, amount: float) -> None:
        potentialNewConcentration = 0
        concentrationIncrease = 0
        indexValue = -1
        for i in range(0,len(self.testTubes)):
            # find ingredients in rack
            if self.testTubes[i].name in ingredient: 
                # calculate the potential new concentration
                potentialNewConcentration = self.testTubes[i].concentrationIfFillToPrice(amount - self.testTubes[i].getCost()) 
                # calculate overall concentration increase
                concentrationIncrease += potentialNewConcentration - self.testTubes[i].concentration 
                indexValue = i
        # check if adding the concentration increase will push the rack out of bounds
        if concentrationIncrease > self.unusedRackConcentration(): 
            # if out of bounds fill to max concnentration
            self.fillToConcentration(ingredient) 
        else:
            # else fill to desired concentration
            self.testTubes[indexValue].addByPrice(amount - self.testTubes[indexValue].getCost()) 



    # Example: Current Tube Concentration = 4, amount = 10, increase tube concentration by 6
    # if increase tube concentration will exceed 100 percent call fil to 100% concentration
    def increaseRackTubeToConcentration(self, ingredient: str, amount: float) -> None:
        for i in range(0, len(self.testTubes)):
            # find ingredients in rack
            if self.testTubes[i].name in ingredient: 
                # check to see if changing to the new amount will push rach concentration out of bounds
                if amount - self.testTubes[i].concentration >= self.unusedRackConcentration(): 
                    # if > 100 send concentration to max
                    self.fillToConcentration([self.testTubes[i].name]) 
                    break
                else:
                    # if < 100 change the concentration to the desired setpoint
                    self.testTubes[i].addByConcentration(amount - self.testTubes[i].concentration) 
                    break



    # changes the concentration of a particular ingredient in the rack tube
    def alterRackTubeConcentration(self, ingredient: str, amount: float) -> None:
        # find the ingredient in question
        for i in range(0, len(self.testTubes)):
            if self.testTubes[i].name in ingredient:
                # decrease the concentration
                if self.testTubes[i].concentration > amount:
                    self.decreaseRackTubeToConcentration(ingredient, amount)
                # increase the concentration
                else:
                    self.increaseRackTubeToConcentration(ingredient, amount)



    # alter the price of an ingredient in the test tube rack
    def alterRackTubePrice(self, ingredient: str, amount: float) -> None:
        # find the ingredient in question
        for i in range(0, len(self.testTubes)):
            if self.testTubes[i].name in ingredient:
                # decrease the price of the ingredient
                if self.testTubes[i].getCost() > amount:
                    self.decreaseRackTubeToPrice(ingredient, amount)
                # increase the price of the ingredient
                else:
                    self.increaseRackTubeToPrice(ingredient, amount)



    # Example: Current Tube Cost = $10, amount = $4, reduce concentration by $6
    # if amount is greater than tube cost or < 0, do nothing
    def decreaseRackTubeToPrice(self, ingredient: str, amount: float) -> None:
        # check amount bounds
        if amount >= 0: 
            for i in range(0, len(self.testTubes)):
                # find ingredients in rack
                if self.testTubes[i].name in ingredient: 
                    # reality check
                    if self.testTubes[i].getCost() > amount: 
                        # reduce price to the amount
                        self.testTubes[i].reduceByPrice(self.testTubes[i].getCost() - amount) 
                        break
                    else:
                        break



    # Example: Current Tube concentration = 10, amount = 4, reduce concentration by 6
    # if amount is greater than tube concentration or < 0, do nothing    
    def decreaseRackTubeToConcentration(self, ingredient: str, amount: float) -> None:
        # check bounds
        if amount >= 0: 
            for i in range(0, len(self.testTubes)):
                # find ingredients in rack
                if self.testTubes[i].name in ingredient: 
                    # reality check
                    if self.testTubes[i].concentration > amount: 
                        # decrease rack tube concentration to amount
                        self.testTubes[i].reduceByConcentration(self.testTubes[i].concentration - amount) 
                        break
                    else:
                        break



    def reduceSolventWhenFillToPricePoint(self, solvents: List[str], ingredientsToChange: List[str]) -> None:
        # empty tubes selected as solvents in the list
        self.emptyTubes(solvents) 
        # fill the ingredients that can change
        self.fillToPrice(ingredientsToChange) 
        # refill the solvents to max concentration
        self.fillToConcentration(solvents) 
        acceptableError = 0.00001
         # Converge to this error, this should be okay for $solvents
        while self.getCost() - self.pricePoint > acceptableError:
            oldCost = self.getCost()
            # reduce the price of the varis evenly
            self.reduceToPrice(ingredientsToChange) 
            # Fill concentration of solvents
            self.fillToConcentration(solvents) 
            newCost = self.getCost()
            if newCost >= oldCost:
                # empty varies
                self.emptyTubes(ingredientsToChange) 
                # empty solvents
                self.emptyTubes(solvents) 
                self.fillToPrice(solvents)
                return



    def increaseSolventWhenReduceToPricePoint(self, solvents: List[str], ingredientsToChange: List[str]) -> None:
        # empty solvents
        self.emptyTubes(solvents) 
        # reduce the ingredients to the price point
        self.reduceToPrice(ingredientsToChange) 
        # fill the solvents
        self.fillToConcentration(solvents) 
        acceptableError = 0.00001
         # Converge to this error, this should be okay for $solvents
        while self.getCost() - self.pricePoint > acceptableError:
            oldCost = self.getCost()
            # reduce the price of the varis evenly
            self.reduceToPrice(ingredientsToChange) 
            # Fill concentration of solvents
            self.fillToConcentration(solvents) 
            newCost = self.getCost()
            if newCost >= oldCost:
                # empty varies
                self.emptyTubes(ingredientsToChange) 
                # empty solvents
                self.emptyTubes(solvents) 
                self.fillToPrice(solvents)
                return



    def batchingInstructions(self, batchSize:float) -> List[Any]:
        # initialize the list of names and batching amount
        ingredientAmount = [] 
        for tt in self.testTubes:
            # append name of ingredient and batching amount
            ingredientAmount.append([tt.name, tt.concentration * batchSize / 100]) 
            print(tt.name + ": " + str(tt.concentration))
        # return names and batching amount * Batching amount is in column 2
        return ingredientAmount 



    def pricePerGallon(self, specificGravity:float) -> float:
        # conversion 1 lb == 453.592 grams
        lbPerGram = 1/453.592 
        # conversion 1 gallon == 3785.41 ml
        mlPerGallon = 3785.41 
        # Calculates the price per gallon. * requires specific gravity input
        return self.getCost() * lbPerGram * specificGravity * mlPerGallon 



#######################################################################################################################################################################

class rackMaker:

    # initializes and manipulates test tube racks
    def createTestTubeRack(self, formulaName: str, formulaLoc: str) -> testTubeRack:
        # opens workbook 
        wb = xlrd.open_workbook(formulaLoc) 
        # get the first sheet in the workbook
        sheet = wb.sheet_by_index(0) 
        # number of rows with content
        i = sheet.nrows
        targetValue = 0
        allTestubes = []
        # get info for each of the ingredients
        for x in range(1, i):
            # creates array based on rows
            value = sheet.row_values(x)
            # handles formulas with notes but no more than one (or two?) ingrdients
            if value[4] == "" and value[5]=="":
                break
            targetValue += value[4] * value[5]
            allTestubes.append(value)
        # calculate target value
        targetValue = targetValue/100
        # create test tube rack
        ttr = testTubeRack(formulaName, targetValue)
        # creates test tubes
        for tt in allTestubes:
            ttr.createRackTube(tt[1], tt[2], tt[5], tt[4])
        try:
            # add notes, if any
            ttr.notes = sheet.cell(rowx = 2, colx = 8).value
        except:
            print("")
        try:
            # add pricepoint, if any
            ttr.pricePoint = float(sheet.cell(rowx = 0, colx = 8).value)
        except:
            print("")
        # return test tube rack
        return ttr



    # TODO: determine why this method was removed?
    # yeah, turns out this is necessary
    # def deleteFormula(self, fileToDelete: str):
    #     dirPath = os.path.dirname(__file__)  
    #     os.chdir(dirPath + "\\Formulas")
    #     try:
    #         os.remove(fileToDelete + ".xlsx")
    #     except:
    #         return



    # Opens excel files
    def openExcelFile(self, fileName: str, oldFileName: str, formula : testTubeRack) -> None:
        # does not need alteration for new formula
        self.saveFormula(fileName, oldFileName, formula)
        # get path to source code
        dirPath = os.path.dirname(__file__)  
        # names the file
        fileName = fileName+".xlsx"
        # TODO: this may need to be fixed or adjusted
        # add to Formulas folder 
        truePath = dirPath+"\\Formulas\\"+fileName
        os.startfile(truePath)
            
    

    def saveFormula(self, fileName: str, oldFileName: str, ttr: testTubeRack) -> None:
        # for new formula, set "oldFileName" to ""
        # no alteration needed for new formula
        # gets path to desktop
        # TODO: get this to make Formula folder in desktop if one doesn't already exist
        dirPath = os.path.join(os.environ["HOMEPATH"], "Desktop")
        # save to Formulas folder
        os.chdir(dirPath + "\\Formulas")
        if oldFileName != "":
            # destroy oldFileName
            os.remove(oldFileName + ".xlsx")
        # write the new file
        ttr.name = fileName
        ttr.exportFormula()


    # formula save as 
    def saveAsFormula(self, fileName: str, oldFileName: str, ttr: testTubeRack) -> str:
        # no alteration needed for new formula
        dirPath = os.path.join(os.environ["HOMEPATH"], "Desktop")
        os.chdir(dirPath + "\\Formulas")
        # if filename exists  in directory
        if path.exists(fileName + ".xlsx"):
            # if the last value of filename is )
            if fileName[-1] == ")":
                # if filename has a (
                if "(" in fileName:
                    openParenth = fileName.rfind("(")
                    potentialFileNumber = fileName[openParenth + 1:-1]
                    # if the content between the open parenthese and close parenthese type(int()) == type(int)
                    try:
                        # then int('number') + 1
                        newInt = int(potentialFileNumber) + 1
                        # TODO: check to see if that save file already exists:
                        # if it does, increment up until you can find a file name that hasn't been used yet
                        newFormulaName = fileName[0:openParenth-1] + " (" + str(newInt) + ")"
                        self.saveFormula(newFormulaName,"", ttr)
                        ttr.name = newFormulaName
                        return newFormulaName
                    except:
                        # save a newFile with the same name but + "(1)"
                        newFormulaName = fileName + " (1)"
                        # TODO: check to see if that save file already exists:
                        # if it does, increment up until you can find a file name that hasn't been used yet
                        self.saveFormula(newFormulaName, "", ttr)
                        ttr.name = newFormulaName
                        return newFormulaName
                else:
                    # save a newFile with the same name but + "(1)"
                    newFormulaName = fileName + " (1)"
                    # TODO: check to see if that save file already exists:
                    # if it does, increment up until you can find a file name that hasn't been used yet
                    self.saveFormula(newFormulaName, "", ttr)
                    ttr.name = newFormulaName
                    return newFormulaName
            else:
                # save a newFile with the same name but + "(1)"
                newFormulaName = fileName + " (1)"
                # TODO: check to see if that save file already exists:
                # if it does, increment up until you can find a file name that hasn't been used yet
                self.saveFormula(newFormulaName, "", ttr)
                ttr.name = newFormulaName
                return newFormulaName
        else:
            # save file without destroying the previous one
            self.saveFormula(fileName, "", ttr)
            ttr.name = fileName
            return fileName